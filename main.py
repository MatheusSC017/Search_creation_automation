"""
Script to automate the process of creating surveys on the ferendum website
"""
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook, Workbook


class RpaFerendum:
    """
    Class to automate the process of creating surveys on the ferendum website, the class will
    receive the path to a file with the surveys and return a new file with the link to the created
    surveys
    """
    def __init__(self, researches, output_file, delimiter=';') -> None:
        """
        this function is the constructor of the RpaFerendum class and defines the initial value for the class variables

        :param researches: Define the path for the .xlsx file that contains the data of the researches that will be
        registered on the Ferendum website
        :param output_file: Set the path to the .xlsx file where the links to the administration and search pages will
        be saved
        :param delimiter: In the options column, each possible answer is separated by a delimiter, the default value is
        ';'
        :return: None
        """
        self._researches = researches
        self._output_file = output_file
        self._delimiter = delimiter
        self._researches_list = []

    @property
    def researches(self) -> str:
        """
        GET method for file with searches
        :return:
        """
        return self._researches

    @researches.setter
    def researches(self, researches) -> None:
        self._researches = researches

    @property
    def output_file(self) -> str:
        """
        GET method for output file name
        :return: string
        """
        return self._output_file

    @output_file.setter
    def output_file(self, output_file) -> None:
        """
        SET method for output file name
        :param output_file: output file name
        :return: None
        """
        self._output_file = output_file

    @property
    def researches_list(self):
        """
        GET method for the output list with research links
        :return: list
        """
        return self._researches_list or []

    def run(self) -> None:
        """ Main function
        create the driver and runs the commands to read the questions and create the researches

        :return: None
        """
        browser = webdriver.Firefox()

        data = self._get_data()

        self._researches_list = []
        for row in data:
            browser.get('https://ferendum.com/pt/')
            self._fill_form(browser, row)
            link_search, link_admin = self._send_form(browser)
            self._researches_list.append((link_search, link_admin))
            time.sleep(1)

    def _get_data(self) -> list:
        """ Get question data
        Reads and extracts the data in the file with the questions

        :return: list of questions
        """
        wb = load_workbook(filename=self._researches)
        sheet = wb['Pesquisas']
        data = []
        for row in sheet:
            row_values = [cell.value for cell in row]
            row_values[3] = [v.strip() for v in row_values[3].split(self._delimiter)]
            row_values[4:] = [0 if v == 'Sim' else 1 for v in row_values[4:]]
            data.append(row_values)
        return data[1:]

    def save(self) -> None:
        """ Save research links
        Save research and admin page links  to an xlsx file

        :return: None
        """
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'Links das pesquisas'

        for row, search in enumerate(self._researches_list, start=1):
            ws1.cell(row=row, column=1, value=search[0])
            ws1.cell(row=row, column=2, value=search[1])

        wb.save(filename=self._output_file)

    @staticmethod
    def _fill_form(driver, data) -> None:
        """ Fill in the form
        Fill in the research form with the question data and the requested settings

        :param driver: the driver to simulate the browser
        :param data: the question data
        :return: None
        """
        driver.find_element(By.NAME, 'titulo').send_keys(data[0])
        driver.find_element(By.NAME, 'descripcion').send_keys(data[1])
        driver.find_element(By.NAME, 'creador').send_keys(data[2])
        for i, op in enumerate(data[3], start=1):
            if i != 1 and i % 5 == 1:
                driver.find_element(By.CLASS_NAME, 'btn-outline-primary').click()
            driver.find_element(By.ID, f'op{i}').send_keys(op)
        driver.find_elements(By.NAME, 'config_anonimo')[data[4]].click()
        driver.find_elements(By.NAME, 'config_priv_pub')[data[5]].click()
        driver.find_elements(By.NAME, 'config_un_solo_voto')[data[6]].click()
        driver.find_elements(By.NAME, 'config_aut_req')[data[7]].click()
        driver.find_element(By.NAME, 'accept_terms_checkbox').click()

    @staticmethod
    def _send_form(driver) -> tuple():
        """ Submit the form
        Submit the research form and get the link to the research and administration page

        :param driver: the driver to simulate the browser
        :return: A tuple with the link to the research and administration page
        """
        driver.find_element(By.CLASS_NAME, 'btn-primary').click()

        time.sleep(1)

        driver.find_element(By.CLASS_NAME, 'btn-primary').click()

        link_search = driver.find_element(By.ID, 'textoACopiar').get_property('href')
        link_admin = driver\
            .find_element(By.CSS_SELECTOR, 'div.card:nth-child(1) > div:nth-child(1) > '
                                           'a:nth-child(22)')\
            .get_property('href')

        return link_search, link_admin


if __name__ == '__main__':
    rpa = RpaFerendum(researches='dados.xlsx', output_file='links.xlsx')
    rpa.run()
    rpa.save()
